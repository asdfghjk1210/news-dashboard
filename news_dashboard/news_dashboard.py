import os
import json
import urllib.parse
from datetime import datetime, timedelta, timezone
from collections import defaultdict

import feedparser
import openpyxl
import yfinance as yf
from dateutil import parser
from openpyxl.styles import Font, PatternFill

# ─────────────────────────────────────
# [1] 설정
# ─────────────────────────────────────
PROVIDER_CONFIG = {
    "Al Jazeera": {"url": "https://www.aljazeera.com/xml/rss/all.xml", "tz": 3, "loc": "국외"},
    "BBC":        {"url": "http://feeds.bbci.co.uk/news/world/rss.xml", "tz": 0, "loc": "국외"},
    "Reuters":    {"url": "https://www.reuters.com/arc/outboundfeeds/rss/topics/world/", "tz": 0, "loc": "국외"},
    "AP":         {"url": "https://apnews.com/rss/world-news", "tz": -5, "loc": "국외"},
    "TASS":       {"url": "https://tass.com/rss/v2.xml", "tz": 3, "loc": "국외"},
    "Kyodo":      {"url": "https://english.kyodonews.net/rss/news.xml", "tz": 9, "loc": "국외"},
    "Xinhua":     {"url": "http://www.xinhuanet.com/english/rss/worldrss.xml", "tz": 8, "loc": "국외"},
    "Google_KR":  {"url": "https://news.google.com/rss/search?q={q}&hl=ko&gl=KR&ceid=KR:ko", "tz": 9, "loc": "국내"}
}

CATEGORIES = {
    "에너지/공급": ["전쟁", "분쟁", "산유국", "OPEC", "가스", "석유", "기름", "oil", "gas", "LNG", "중동", "이란", "이스라엘"],
    "에너지/인프라": ["발전소", "전력", "전기", "전기료", "에너지", "energy", "항만", "송전"],
    "시장/금융": ["금리", "환율", "주식", "채권", "물가", "인플레이션", "interest", "stock", "exchange", "inflation", "연준", "Fed"],
    "생활/민생": ["전기료", "택시", "배달", "물류", "화물차", "민생", "요금", "장바구니"],
    "정책/대응": ["지원금", "보조금", "대책", "규제", "예산", "지자체", "policy", "subsidy", "긴급", "보증", "특례"]
}

GU_KEYWORDS = {
    "팔달구": ["팔달", "수원 전통시장", "수원 남문", "소상공인"],
    "장안구": ["장안", "수원 북", "농촌", "농업"],
    "영통구": ["영통", "광교", "삼성", "R&D", "전기차"],
    "권선구": ["권선", "수원 산단", "공단", "제조"],
}

YOUTUBE_CHANNELS = {
    "경제·금융 지식": [
        {"name": "슈카월드", "id": "UCsJ6RuBiTVWRX156FVbeaGg", "desc": "중동 사태, 국제 유가 전망 · 시장 금융 영향 분석"},
        {"name": "삼프로TV",  "id": "UChlv4GSd7OQl3js-jkLOnFA", "desc": "금리/환율, 개인 투자 전략 · 부동산 시장 대응"},
    ],
    "주식 및 테크": [
        {"name": "염블리(염승환)", "id": "UCDnIfMiHBNs1RP7y6B6wf1Q", "desc": "국내외 주식 종목 분석 · 투자 전략"},
        {"name": "월급쟁이 부자들", "id": "UCDSj40X9FFUAnx1nv7gQhcA", "desc": "자산 형성 전략 · 부동산 투자 대응"},
    ],
    "자산 및 금리": [
        {"name": "이효석의 입풀", "id": "UCxvdCnvGODDyuvnELnLkQWw", "desc": "금리 동향 · 자산 시장 구조 분석"},
    ],
    "핵심 브리핑": [
        {"name": "소수몽키",    "id": "UCC3yfxS5qC6PCwDzetUuEWg", "desc": "모든 만에 보는 핵심 이슈 · 주간 경제 동향"},
        {"name": "경제갤러리",  "id": "UCuPWXwE1pnjhNIZ2AStzOVA", "desc": "시장 갤러리 브리핑 · 투자자 심리 분석"},
        {"name": "유쾌한경제학","id": "UCaLsqraxQaaRfm7HyT1x6ug", "desc": "쉽게 읽는 경제 뉴스 · 정책 해설"},
    ]
}

OUTPUT_XLSX = "news_result.xlsx"
OUTPUT_HTML = "docs/market.html"
DAYS_RANGE  = 7

# ─────────────────────────────────────
# [2] 분류 함수
# ─────────────────────────────────────

def classify_category(title):
    title_lower = title.lower()
    for cat, kws in CATEGORIES.items():
        if any(kw.lower() in title_lower for kw in kws):
            return cat
    return "기타"

def classify_gu(title):
    title_lower = title.lower()
    for gu, kws in GU_KEYWORDS.items():
        if any(kw in title_lower for kw in kws):
            return gu
    return None

def build_weekly_chart_data(articles):
    kst = timezone(timedelta(hours=9))
    today = datetime.now(kst).date()
    dates = [(today - timedelta(days=6-i)) for i in range(7)]
    date_labels = [d.strftime("%b %d") for d in dates]

    cat_counts = {cat: {d: 0 for d in dates} for cat in CATEGORIES}
    for a in articles:
        pub_date = a["발행KST"].date()
        cat = a["분류"]
        if cat in cat_counts and pub_date in cat_counts[cat]:
            cat_counts[cat][pub_date] += 1

    COLORS = {
        "에너지/공급":   "#3b82f6",
        "에너지/인프라": "#06b6d4",
        "시장/금융":     "#10b981",
        "생활/민생":     "#f59e0b",
        "정책/대응":     "#ef4444",
    }
    datasets = []
    for cat, color in COLORS.items():
        datasets.append({
            "label": cat,
            "data": [cat_counts[cat][d] for d in dates],
            "borderColor": color,
            "backgroundColor": color + "14",
            "tension": 0.4,
            "pointRadius": 3,
            "borderWidth": 2
        })
    return date_labels, datasets

def build_gu_summaries(articles):
    gu_data = {gu: defaultdict(list) for gu in GU_KEYWORDS}
    for a in articles:
        gu = classify_gu(a["제목"])
        if gu:
            gu_data[gu][a["분류"]].append(a["제목"])

    top_by_cat = defaultdict(list)
    for a in articles[:30]:
        top_by_cat[a["분류"]].append(a["제목"])

    GU_COLORS = {"팔달구": "#3b82f6", "장안구": "#10b981", "영통구": "#f59e0b", "권선구": "#ef4444"}
    result = {}
    for gu, color in GU_COLORS.items():
        cats = gu_data[gu]
        lines = []
        shown_cats = set()
        for cat, titles in cats.items():
            if cat == "기타": continue
            lines.append({"cat": cat, "items": [t[:38] + ("…" if len(t) > 38 else "") for t in titles[:2]]})
            shown_cats.add(cat)
            if len(lines) >= 3: break
        if len(lines) < 2:
            for cat, titles in top_by_cat.items():
                if cat not in shown_cats and cat != "기타" and titles:
                    lines.append({"cat": cat, "items": [titles[0][:38] + "…"]})
                    if len(lines) >= 3: break
        result[gu] = {"color": color, "lines": lines}
    return result

def build_policy_summary(articles):
    return {
        "에너지": [a["제목"][:42] + "…" for a in articles if a["분류"] in ("에너지/공급", "에너지/인프라")][:2],
        "민생":   [a["제목"][:42] + "…" for a in articles if a["분류"] == "생활/민생"][:2],
        "산업":   [a["제목"][:42] + "…" for a in articles if a["분류"] == "시장/금융"][:2],
        "정책":   [a["제목"][:42] + "…" for a in articles if a["분류"] == "정책/대응"][:2],
    }

# ─────────────────────────────────────
# [3] 데이터 수집
# ─────────────────────────────────────

def get_econ_indicators():
    print("  - 경제 지표 수집 중...")
    try:
        tickers = {"환율(USD/KRW)": "USDKRW=X", "금값": "GC=F", "WTI유가": "CL=F"}
        import yfinance as yf
        data = yf.download(list(tickers.values()), period="1mo", interval="1d", progress=False)["Close"]
        data.columns = list(tickers.keys())
        tail = data.dropna().tail(10)
        return tail.to_dict("list"), tail.index.strftime("%m-%d").tolist()
    except Exception as e:
        print(f"    경제 지표 오류: {e}")
        return {}, []

def collect_news():
    articles = []
    now_utc = datetime.now(timezone.utc)
    cutoff  = now_utc - timedelta(days=DAYS_RANGE)
    kst     = timezone(timedelta(hours=9))
    print("  - 국내외 뉴스 통합 수집 중...")
    for name, config in PROVIDER_CONFIG.items():
        url = config["url"]
        if name == "Google_KR":
            q   = urllib.parse.quote("수원시 OR 에너지 OR 가스 OR 금리 OR 물가 OR 중동 OR 유가")
            url = url.format(q=q)
        try:
            feed = feedparser.parse(url)
        except Exception as e:
            print(f"    {name} 피드 오류: {e}")
            continue
        for entry in feed.entries:
            try:
                raw_date = parser.parse(entry.published)
                if raw_date.tzinfo is None:
                    raw_date = raw_date.replace(tzinfo=timezone(timedelta(hours=config["tz"])))
                pub_utc = raw_date.astimezone(timezone.utc)
                if pub_utc < cutoff: continue
                articles.append({
                    "지역":    config["loc"],
                    "언론사":  name,
                    "제목":    entry.title,
                    "링크":    entry.link,
                    "분류":    classify_category(entry.title),
                    "발행UTC": pub_utc,
                    "수집UTC": datetime.now(timezone.utc),
                    "발행KST": pub_utc.astimezone(kst),
                })
            except:
                continue
    return sorted(articles, key=lambda x: x["발행UTC"], reverse=True)

# ─────────────────────────────────────
# [4] HTML 생성
# ─────────────────────────────────────

def save_all(articles, econ_data, econ_dates):
    os.makedirs("docs", exist_ok=True)

    # 엑셀
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monitoring_Data"
    header = ["지역","분류","언론사","기사제목","발행시간(KST)","시스템수집시간(UTC)","원문링크"]
    ws.append(header)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
    for a in articles:
        ws.append([a["지역"],a["분류"],a["언론사"],a["제목"],
                   a["발행KST"].strftime("%Y-%m-%d %H:%M"),
                   a["수집UTC"].strftime("%Y-%m-%d %H:%M"),a["링크"]])
    wb.save(OUTPUT_XLSX)

    # 분석
    date_labels, kw_datasets = build_weekly_chart_data(articles)
    gu_summaries             = build_gu_summaries(articles)
    policy                   = build_policy_summary(articles)

    # 구별 HTML
    GU_BORDER = {"팔달구":"#3b82f6","장안구":"#10b981","영통구":"#f59e0b","권선구":"#ef4444"}
    GU_TEXT   = {"팔달구":"#1e40af","장안구":"#065f46","영통구":"#92400e","권선구":"#991b1b"}
    gu_html = ""
    for gu, color in GU_BORDER.items():
        lines = gu_summaries.get(gu, {}).get("lines", [])
        inner = ""
        for line in lines:
            items_html = "".join([f'<div style="font-size:9.5px;color:#475569;line-height:1.5">· {item}</div>' for item in line["items"]])
            inner += f'<div style="margin-bottom:4px"><span style="font-size:9px;font-weight:500;color:{color}">[{line["cat"]}]</span><br>{items_html}</div>'
        if not inner:
            inner = '<span style="font-size:9.5px;color:#94a3b8">수집된 기사 없음</span>'
        gu_html += f'<div style="border-left:3px solid {color};padding:4px 8px;margin-bottom:6px;background:#fafafa;border-radius:0 4px 4px 0"><div style="font-size:11px;font-weight:500;color:{GU_TEXT[gu]};margin-bottom:4px">{gu}</div>{inner}</div>'

    # 유튜브 탭 HTML
    yt_buttons = ""
    yt_contents = ""
    for i, (group, channels) in enumerate(YOUTUBE_CHANNELS.items()):
        active  = "active" if i == 0 else ""
        display = "block"  if i == 0 else "none"
        yt_buttons += f'<button class="ytbtn {active}" onclick="openTab(event,\'yt{i}\')">{group}</button>'
        items = ""
        for ch in channels:
            items += f'<div style="display:flex;align-items:center;gap:6px;padding:5px 6px;background:#f8fafc;border-radius:4px;border:0.5px solid #e2e8f0;margin-bottom:4px"><iframe width="80" height="50" src="https://www.youtube.com/embed?listType=user_uploads&list={ch["id"]}" frameborder="0" allowfullscreen style="border-radius:3px;flex-shrink:0"></iframe><div><div style="font-size:10px;font-weight:500;color:#1e40af">{ch["name"]}</div><div style="font-size:9px;color:#64748b;line-height:1.4">{ch["desc"]}</div></div></div>'
        yt_contents += f'<div id="yt{i}" style="display:{display}">{items}</div>'

    # 정책 HTML
    colors_p  = {"에너지":"#10b981","민생":"#3b82f6","산업":"#f59e0b","정책":"#8b5cf6"}
    tcolors_p = {"에너지":"#065f46","민생":"#1e40af","산업":"#92400e","정책":"#5b21b6"}
    policy_html = ""
    for label, items in policy.items():
        c  = colors_p.get(label,"#94a3b8")
        tc = tcolors_p.get(label,"#374151")
        items_html = "".join([f'<div style="font-size:9px;color:#475569;line-height:1.5">· {i}</div>' for i in items]) or '<div style="font-size:9px;color:#94a3b8">수집된 기사 없음</div>'
        policy_html += f'<div style="border-left:3px solid {c};padding:4px 8px;margin-bottom:6px"><div style="font-size:10px;font-weight:500;color:{tc};margin-bottom:2px">[{label}]</div>{items_html}</div>'

    # 뉴스 테이블
    cat_colors = {"에너지/공급":"#0ea5e9","에너지/인프라":"#06b6d4","시장/금융":"#10b981","생활/민생":"#f59e0b","정책/대응":"#8b5cf6","기타":"#94a3b8"}
    news_rows = ""
    for a in articles[:50]:
        bc = "#3b82f6" if a["지역"]=="국내" else "#ef4444"
        cc = cat_colors.get(a["분류"],"#94a3b8")
        news_rows += f'<tr><td><span style="background:{bc};color:white;padding:2px 6px;border-radius:4px;font-size:9px">{a["지역"]}</span></td><td><span style="background:{cc};color:white;padding:2px 6px;border-radius:4px;font-size:9px">{a["분류"]}</span></td><td style="color:#374151">{a["언론사"]}</td><td style="color:#64748b">{a["발행KST"].strftime("%m-%d %H:%M")}</td><td><a href="{a["링크"]}" target="_blank" style="color:#3b82f6;text-decoration:none">{a["제목"][:55]}{"…" if len(a["제목"])>55 else ""}</a></td></tr>'

    # 종합 현황
    cat_counts_total = defaultdict(int)
    for a in articles:
        cat_counts_total[a["분류"]] += 1
    top_cats = sorted(cat_counts_total.items(), key=lambda x: x[1], reverse=True)[:3]
    issues_html = "".join([f'<div style="margin-bottom:4px;font-size:10px;color:#374151">· [{cat}] 관련 기사 <strong>{cnt}건</strong> — 주요 이슈 집중 모니터링 필요</div>' for cat, cnt in top_cats])

    # top5
    top5_html = "".join([f'<li style="margin-bottom:5px;font-size:11px"><span style="background:#ef4444;color:white;padding:1px 5px;border-radius:3px;font-size:9px;margin-right:4px">{a["언론사"]}</span><a href="{a["링크"]}" target="_blank" style="color:#1e40af;text-decoration:none">{a["제목"][:60]}{"…" if len(a["제목"])>60 else ""}</a></li>' for a in articles[:5]])

    # 경제 지표 차트
    econ_js = f"""new Chart(document.getElementById('econChart'), {{
        type:'line',
        data:{{
            labels:{json.dumps(econ_dates)},
            datasets:[
                {{label:'환율(USD/KRW)',data:{json.dumps([round(v,1) if v else None for v in econ_data.get('환율(USD/KRW)',[])])},borderColor:'#3b82f6',tension:0.3,pointRadius:3,borderWidth:2,yAxisID:'y'}},
                {{label:'WTI유가',data:{json.dumps([round(v,2) if v else None for v in econ_data.get('WTI유가',[])])},borderColor:'#ef4444',tension:0.3,pointRadius:3,borderWidth:2,yAxisID:'y2'}},
                {{label:'금값',data:{json.dumps([round(v,1) if v else None for v in econ_data.get('금값',[])])},borderColor:'#f59e0b',tension:0.3,pointRadius:3,borderWidth:2,yAxisID:'y2'}}
            ]
        }},
        options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{labels:{{font:{{size:10}},boxWidth:12}}}}}},scales:{{x:{{ticks:{{font:{{size:9}}}},grid:{{display:false}}}},y:{{ticks:{{font:{{size:9}}}}}},y2:{{position:'right',ticks:{{font:{{size:9}}}},grid:{{drawOnChartArea:false}}}}}}}}
    }});"""

    now_str = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d %H:%M KST")

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>수원시 구별 중동 상황 및 분석</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Malgun Gothic','Apple SD Gothic Neo',sans-serif;background:#f0f2f5;color:#1e293b;font-size:12px}}
.wrap{{max-width:1400px;margin:0 auto;padding:10px}}
.main-hdr{{background:#1e3a8a;color:white;padding:10px 16px;border-radius:8px 8px 0 0;display:flex;justify-content:space-between;align-items:center}}
.main-hdr h1{{font-size:15px;font-weight:500}}
.nav-tabs{{background:#1e40af;display:flex;gap:2px;padding:0 16px;margin-bottom:8px;border-radius:0 0 4px 4px}}
.nav-tabs a{{color:rgba(255,255,255,.7);text-decoration:none;padding:7px 14px;font-size:11px;border-bottom:2px solid transparent}}
.nav-tabs a.active{{color:white;border-bottom:2px solid white}}
.row{{display:grid;gap:8px;margin-bottom:8px}}
.r3{{grid-template-columns:1.1fr 1.6fr 1.3fr}}
.r2{{grid-template-columns:1fr 1fr}}
.r3b{{grid-template-columns:1fr 1fr 1fr}}
.card{{background:white;border-radius:8px;border:0.5px solid #e2e8f0;padding:10px;overflow:hidden}}
.card-hdr{{font-size:11px;font-weight:500;color:#1e3a8a;border-bottom:1px solid #e2e8f0;padding-bottom:6px;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center}}
.sec-wrap{{border-radius:8px;overflow:hidden;border:0.5px solid #e2e8f0}}
.sec-hdr{{background:#1e3a8a;color:white;padding:7px 12px;font-size:11px;font-weight:500;display:flex;justify-content:space-between;align-items:center}}
.sec-body{{background:white;padding:10px}}
.ytbtn{{padding:3px 9px;border:0.5px solid #cbd5e1;background:white;border-radius:4px;font-size:9.5px;cursor:pointer;color:#475569;margin-right:3px;margin-bottom:4px}}
.ytbtn.active{{background:#1e3a8a;color:white;border-color:#1e3a8a}}
table{{width:100%;border-collapse:collapse;font-size:10px}}
th{{background:#1e3a8a;color:white;padding:6px 8px;text-align:left;font-weight:500}}
td{{padding:5px 8px;border-bottom:0.5px solid #f1f5f9;vertical-align:middle}}
tr:hover td{{background:#f8fafc}}
.dl-btn{{background:#10b981;color:white;padding:4px 12px;border-radius:4px;border:none;font-size:10px;cursor:pointer;font-weight:500;text-decoration:none;display:inline-block}}
.chart-wrap{{height:155px;position:relative}}
.region-item{{font-size:9px;padding:3px 6px;margin-bottom:3px;border-radius:0 3px 3px 0;line-height:1.4}}
</style>
</head>
<body>
<div class="wrap">

  <div class="main-hdr">
    <h1>📈 수원시 구별 중동 상황 및 분석 대시보드</h1>
    <span style="font-size:10px;opacity:.75">마지막 업데이트: {now_str}</span>
  </div>
  <div class="nav-tabs">
    <a href="#" class="active">Overview</a>
    <a href="#">Intables</a>
    <a href="#">Githuns</a>
    <a href="#">Preccity</a>
    <a href="#">Carinsights</a>
  </div>

  <!-- 상단 3열 -->
  <div class="row r3">
    <div class="card">
      <div class="card-hdr">수원시 구별 상황 유형 <span style="font-size:9px;color:#64748b">RSS 자동 분류</span></div>
      {gu_html}
    </div>
    <div class="card">
      <div class="card-hdr">주간 유형별 키워드 검색량 (뉴스 카운팅 기반)</div>
      <div class="chart-wrap"><canvas id="kwChart"></canvas></div>
      <div class="card-hdr" style="margin-top:8px">주요 경제 지표 (USD/KRW · WTI · 금값)</div>
      <div class="chart-wrap"><canvas id="econChart"></canvas></div>
    </div>
    <div class="card">
      <div class="card-hdr">전문가 채널 브리핑</div>
      <div style="margin-bottom:6px">{yt_buttons}</div>
      {yt_contents}
    </div>
  </div>

  <!-- 중단 2열 -->
  <div class="row r2">
    <div class="sec-wrap">
      <div class="sec-hdr">종합 상황 및 정책 제안 <a href="{OUTPUT_XLSX}" class="dl-btn">종합 보고서 (xlsx) 다운로드</a></div>
      <div class="sec-body">
        <div style="font-size:10px;font-weight:500;color:#374151;margin-bottom:6px">▶ 수원시 종합 현황 및 문제점 (RSS 자동 집계)</div>
        {issues_html}
        <div style="font-size:10px;font-weight:500;color:#374151;margin:10px 0 6px">▶ 통합 대응 정책 제안 (기사 기반 자동 요약)</div>
        {policy_html}
      </div>
    </div>
    <div class="sec-wrap">
      <div class="sec-hdr">다각도 상황 및 대응 현황</div>
      <div class="sec-body">
        <div class="row r3b">
          <div>
            <div style="font-size:10px;font-weight:500;color:#1e40af;border-bottom:1px solid #e2e8f0;padding-bottom:4px;margin-bottom:6px">국외 현황</div>
            {''.join([f'<div class="region-item" style="background:#eff6ff;border-left:2px solid #3b82f6"><a href="{a["링크"]}" target="_blank" style="color:#1e40af;text-decoration:none">{a["제목"][:38]}{"…" if len(a["제목"])>38 else ""}</a></div>' for a in articles if a["지역"]=="국외"][:6])}
          </div>
          <div>
            <div style="font-size:10px;font-weight:500;color:#065f46;border-bottom:1px solid #e2e8f0;padding-bottom:4px;margin-bottom:6px">국내 현황</div>
            {''.join([f'<div class="region-item" style="background:#f0fdf4;border-left:2px solid #10b981"><a href="{a["링크"]}" target="_blank" style="color:#065f46;text-decoration:none">{a["제목"][:38]}{"…" if len(a["제목"])>38 else ""}</a></div>' for a in articles if a["지역"]=="국내"][:6])}
          </div>
          <div>
            <div style="font-size:10px;font-weight:500;color:#92400e;border-bottom:1px solid #e2e8f0;padding-bottom:4px;margin-bottom:6px">실시간 Top 5</div>
            <ol style="padding-left:14px">{top5_html}</ol>
          </div>
        </div>
      </div>
    </div>
  </div>

  <!-- 하단 테이블 -->
  <div class="sec-wrap">
    <div class="sec-hdr">국내외 언론 보도 모니터링 데이터 <a href="{OUTPUT_XLSX}" class="dl-btn">크롤링 데이터 (엑셀.xlsx) 다운로드</a></div>
    <div style="overflow-x:auto">
      <table>
        <thead><tr><th>지역</th><th>분류</th><th>언론사</th><th>보도일(KST)</th><th>제목 (클릭 시 원문)</th></tr></thead>
        <tbody>{news_rows}</tbody>
      </table>
    </div>
  </div>

</div>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script>
new Chart(document.getElementById('kwChart'), {{
  type:'line',
  data:{{labels:{json.dumps(date_labels)},datasets:{json.dumps(kw_datasets)}}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{labels:{{font:{{size:9}},boxWidth:10}}}}}},scales:{{x:{{ticks:{{font:{{size:9}}}},grid:{{display:false}}}},y:{{ticks:{{font:{{size:9}}}},beginAtZero:true}}}}}}
}});
{econ_js}
function openTab(evt,id){{
  document.querySelectorAll('[id^="yt"]').forEach(el=>el.style.display='none');
  document.querySelectorAll('.ytbtn').forEach(btn=>btn.classList.remove('active'));
  document.getElementById(id).style.display='block';
  evt.currentTarget.classList.add('active');
}}
</script>
</body>
</html>"""

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  - HTML 저장 완료: {OUTPUT_HTML}")

# ─────────────────────────────────────
# [5] 실행
# ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  수원시 중동 상황 대시보드 파이프라인 가동")
    print("=" * 55)
    econ_data, econ_dates = get_econ_indicators()
    all_news = collect_news()
    print(f"  - 수집된 기사: {len(all_news)}건")
    save_all(all_news, econ_data, econ_dates)
    print("✅  완료! docs/market.html 을 확인하세요.")
    print("=" * 55)
