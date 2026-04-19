import sys
import os
import json
import urllib.parse
import time
from datetime import datetime, timedelta, timezone

import requests
import feedparser
import openpyxl
import yfinance as yf
import pandas as pd
from bs4 import BeautifulSoup
from dateutil import parser
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────
# [1] 설정 및 매체별 타임존 매핑 테이블
# ─────────────────────────────────────
PROVIDER_CONFIG = {
    "Al Jazeera": {"url": "https://www.aljazeera.com/xml/rss/all.xml", "tz": 3, "loc": "국외"},
    "BBC": {"url": "http://feeds.bbci.co.uk/news/world/rss.xml", "tz": 0, "loc": "국외"},
    "Reuters": {"url": "https://www.reuters.com/arc/outboundfeeds/rss/topics/world/", "tz": 0, "loc": "국외"},
    "AP": {"url": "https://apnews.com/rss/world-news", "tz": -5, "loc": "국외"},
    "TASS": {"url": "https://tass.com/rss/v2.xml", "tz": 3, "loc": "국외"},
    "Kyodo": {"url": "https://english.kyodonews.net/rss/news.xml", "tz": 9, "loc": "국외"},
    "Xinhua": {"url": "http://www.xinhuanet.com/english/rss/worldrss.xml", "tz": 8, "loc": "국외"},
    "Google_KR": {"url": "https://news.google.com/rss/search?q={q}&hl=ko&gl=KR&ceid=KR:ko", "tz": 9, "loc": "국내"}
}

CATEGORIES = {
    "원인/공급 요인": ["전쟁", "분쟁", "산유국", "OPEC", "가스", "기름 폭등", "석유", "석유 폭등", "oil", "gas", "oil price"],
    "에너지/인프라": ["발전소", "전력", "전기", "전기료", "전기 요금", "요금 인상", "요금 폭등", "에너지", "energy", "항만", "LNG", "energy crisis"],
    "시장/금융": ["금리", "환율", "주식", "채권", "물가", "인플레이션", "스테그플레이션", "interest", "stock", "exchange", "inflation", "Stagflation", "rate hike"],
    "생활/영향": ["전기료", "택시", "배달", "물류", "화물차", "민생", "요금", "Electricity bill", "Taxi", "Delivery", "Logistics", "Truck"],
    "정책/대응": ["지원금", "보조금", "대책", "규제", "예산", "지자체", "policy", "response", "subsidy", "tariff", "utility bill"]
}

# 수원시 구별 상황 유형 데이터
SUWON_GU_DATA = {
    "팔달구": [
        "에너지 및 금곰 요인: 주요 가스/기름 안상 추이, 할당구 내 중동 소재 고취 현황",
        "시장 및 금융 요인: 글로벌 소상공인 내환 금리 부분 평가, 팔달 가이스 계정 (미상/배경 제시)",
        "생활 및 민감 요인: 미디어/역략물 안상, 물가 통계 자체 (미상)",
        "정책 대응: 팔달구 소상공인 지원 현황 직접 지원 현황"
    ],
    "장안구": [
        "에너지 및 금곰 요인: 국정 소비임상 할인 및 가격 인상 방향 (비상/배경 계획)",
        "금융 및 인상 요인: 물가 스타 혁명 구 및 대중소기업 연계 영향",
        "생활 및 민감 요인: 물스 스타 혁명 구 및 대중소기업 연계 영향",
        "정책 대응: 장안구 소상공인 지원 연계 영향"
    ],
    "영통구": [
        "에너지 및 금곰 요인: 영통구 나 대정원의 조 기름 활활, 전기 출범 인프라 (산업)",
        "금융 및 인상 요인: 전기/기름 가격 연계 생산 방향",
        "생활 및 민감 요인: 생활물기 연계 생산 영향",
        "정책 대응: 영통구 스타트업 R&D 자원 연계 생산 영향"
    ],
    "권선구": [
        "에너지 및 금곰 요인: 산업단지 이봉민 처에, 물류 보고 내용",
        "생활 및 민감 요인: 화물분포 이상, 물류거 내역",
        "정책 대응: 권선구 소상공인 지원 보조자 지원"
    ]
}

# 유튜브 채널 분류 데이터
YOUTUBE_CHANNELS = {
    "경제·금융 지식": [
        {"name": "슈카월드", "id": "UCsJ6RuBiTVWRX156FVbeaGg"},
        {"name": "삼프로TV", "id": "UChlv4GSd7OQl3js-jkLOnFA"},
    ],
    "주식 및 테크": [
        {"name": "염블리(염승환)", "id": "UCDnIfMiHBNs1RP7y6B6wf1Q"},
        {"name": "월급쟁이 부자들", "id": "UCDSj40X9FFUAnx1nv7gQhcA"}
    ],
    "자산 및 금리": [
        {"name": "이효석의 입풀", "id": "UCxvdCnvGODDyuvnELnLkQWw"},
    ],
    "핵심 브리핑": [
        {"name": "소수몽키", "id": "UCC3yfxS5qC6PCwDzetUuEWg"},
        {"name": "경제갤러리", "id": "UCuPWXwE1pnjhNIZ2AStzOVA"},
        {"name": "유쾌한경제학", "id": "UCaLsqraxQaaRfm7HyT1x6ug"}
    ]
}

# 전문가 채널 브리핑 카드 데이터 (상단 패널용)
EXPERT_BRIEFING = {
    "경제 지식": [
        {
            "channels": "홍춘욱 시니, 국채 유가 진행",
            "desc": "시장 시 에, 국채 유가 전달 | 시장 금융 영향 분석",
            "duration": "3:42"
        },
        {
            "channels": "슈카월드, 삼프로TV",
            "desc": "국제 경제 흐름 및 글로벌 시장 종합 분석",
            "duration": "5:18"
        }
    ],
    "금융 지식": [
        {
            "channels": "염블리(염승환), 월급쟁이 부자들",
            "desc": "금리/환율, 개인 투자 전략 | 무동산 시장 이해",
            "duration": "4:05"
        }
    ],
    "핵심 이슈": [
        {
            "channels": "핵심브리핑: 소수몽키, 경제갤러리",
            "desc": "오늘 안에 보는 핵심 정리 | 주간 경제 통합 정리",
            "duration": "2:38"
        }
    ]
}

OUTPUT_XLSX = "news_result.xlsx"
OUTPUT_HTML = "docs/market.html"
DAYS_RANGE = 3

# ─────────────────────────────────────
# [2] 기능 함수부
# ─────────────────────────────────────

def classify_category(title):
    title_lower = title.lower()
    for cat, kws in CATEGORIES.items():
        if any(kw in title_lower for kw in kws):
            return cat
    return "기타"

def get_econ_indicators():
    print("  - 경제 지표 수집 중...")
    try:
        tickers = {"환율(USD/KRW)": "USDKRW=X", "금값": "GC=F", "WTI유가": "CL=F"}
        data = yf.download(list(tickers.values()), period="1mo", interval="1d")['Close']
        data.columns = [list(tickers.keys())[list(tickers.values()).index(col)] for col in data.columns]
        return data.dropna().tail(10).to_dict('list'), data.index.strftime('%m-%d').tolist()
    except:
        return {}, []

def collect_news():
    articles = []
    now_utc = datetime.now(timezone.utc)
    cutoff = now_utc - timedelta(days=DAYS_RANGE)
    kst = timezone(timedelta(hours=9))

    print("  - 국내외 뉴스 통합 수집 및 시차 교정 중...")
    for name, config in PROVIDER_CONFIG.items():
        url = config["url"]
        if name == "Google_KR":
            q = urllib.parse.quote("수원시 OR 가스 OR 금리 OR 정책")
            url = url.format(q=q)
        feed = feedparser.parse(url)
        for entry in feed.entries:
            try:
                raw_date = parser.parse(entry.published)
                if raw_date.tzinfo is None:
                    raw_date = raw_date.replace(tzinfo=timezone(timedelta(hours=config["tz"])))
                pub_utc = raw_date.astimezone(timezone.utc)
                if pub_utc < cutoff: continue
                crawled_utc = datetime.now(timezone.utc)
                articles.append({
                    "지역": config["loc"], "언론사": name, "제목": entry.title, "링크": entry.link,
                    "분류": classify_category(entry.title), "발행UTC": pub_utc,
                    "수집UTC": crawled_utc, "발행KST": pub_utc.astimezone(kst)
                })
            except:
                continue
    return sorted(articles, key=lambda x: x['발행UTC'], reverse=True)

# ─────────────────────────────────────
# [3] 출력 및 저장부
# ─────────────────────────────────────

def build_suwon_panel():
    """수원시 구별 상황 유형 패널 HTML 생성"""
    items_html = ""
    for gu, details in SUWON_GU_DATA.items():
        detail_lines = "".join([f"<li>{d}</li>" for d in details])
        items_html += f"""
        <div class="gu-item">
            <div class="gu-header" onclick="toggleGu(this)">
                <span class="gu-arrow">▶</span>
                <span class="gu-name">{gu}</span>
            </div>
            <ul class="gu-detail" style="display:none;">
                {detail_lines}
            </ul>
        </div>"""
    return items_html

def build_expert_briefing_panel():
    """전문가 채널 브리핑 카드 HTML 생성"""
    cards_html = ""
    for category, items in EXPERT_BRIEFING.items():
        section_label = f'<div class="briefing-section-label">▌ {category}</div>'
        cards_html += section_label
        for item in items:
            cards_html += f"""
            <div class="briefing-card">
                <div class="briefing-yt-icon">▶</div>
                <div class="briefing-info">
                    <div class="briefing-channels">{item['channels']}</div>
                    <div class="briefing-desc">{item['desc']}</div>
                </div>
                <div class="briefing-duration">{item['duration']}</div>
            </div>"""
    return cards_html

def build_youtube_tabs():
    """유튜브 탭 HTML 생성"""
    tab_buttons = ""
    tab_contents = ""
    for i, (group, channels) in enumerate(YOUTUBE_CHANNELS.items()):
        active_class = "active" if i == 0 else ""
        display_style = "display:block" if i == 0 else "display:none"
        tab_buttons += f'<button class="tablinks {active_class}" onclick="openYoutubeTab(event, \'tab{i}\')">{group}</button>'
        video_grid = '<div class="video-grid">'
        for ch in channels:
            video_grid += f"""
            <div class="video-item">
                <small>{ch["name"]}</small>
                <iframe width="100%" height="180"
                    src="https://www.youtube.com/embed?listType=user_uploads&list={ch["id"]}"
                    frameborder="0" allowfullscreen></iframe>
            </div>"""
        video_grid += '</div>'
        tab_contents += f'<div id="tab{i}" class="tabcontent" style="{display_style}">{video_grid}</div>'
    return tab_buttons, tab_contents

def save_all(articles, econ_data, econ_dates):
    # ── 엑셀 저장 ──────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monitoring_Data"
    header = ["지역", "분류", "언론사", "기사제목", "발행시간(KST)", "시스템수집시간(UTC)", "원문링크"]
    ws.append(header)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(color="FFFFFF", bold=True)
    for a in articles:
        ws.append([
            a["지역"], a["분류"], a["언론사"], a["제목"],
            a["발행KST"].strftime("%Y-%m-%d %H:%M"),
            a["수집UTC"].strftime("%Y-%m-%d %H:%M"),
            a["링크"]
        ])
    wb.save(OUTPUT_XLSX)

    # ── 패널 HTML 조각 생성 ────────────────────────────
    suwon_panel_html    = build_suwon_panel()
    briefing_panel_html = build_expert_briefing_panel()
    tab_buttons, tab_contents = build_youtube_tabs()

    # ── 뉴스 HTML 조각 ─────────────────────────────────
    top5_html  = "".join([f"<li><b>[{a['언론사']}]</b> {a['제목']}</li>" for a in articles[:5]])
    news_rows  = "".join([
        f"<tr>"
        f"<td>{a['발행KST'].strftime('%m-%d %H:%M')}</td>"
        f"<td><span class='badge {a['지역']}'>{a['지역']}</span></td>"
        f"<td>{a['분류']}</td>"
        f"<td>{a['언론사']}</td>"
        f"<td><a href='{a['링크']}' target='_blank'>{a['제목']}</a></td>"
        f"</tr>"
        for a in articles
    ])

    # ── 키워드 차트 데이터 (날짜 레이블은 econ_dates 재활용, 없으면 기본값) ──
    chart_labels = econ_dates[-7:] if len(econ_dates) >= 7 else ["Apr 14","Apr 15","Apr 16","Apr 17","Apr 18","Apr 19","Apr 20"]

    # ── 전체 HTML ──────────────────────────────────────
    html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
    <meta charset="utf-8">
    <title>통합 마켓 인사이트</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        /* ── 기본 레이아웃 ── */
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: 'Malgun Gothic', sans-serif;
            background: #0d1520;
            color: #cdd6e0;
            margin: 20px;
        }}
        .container {{ max-width: 1500px; margin: auto; }}

        h1 {{ color: #e2e8f0; font-size: 22px; }}

        /* ── 2열 메인 그리드 ── */
        .grid-2col {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 16px;
            margin-bottom: 16px;
        }}

        /* ── 카드 ── */
        .card {{
            background: #152032;
            padding: 16px;
            border-radius: 10px;
            border: 1px solid #1e3a5f;
        }}
        .card h3 {{
            color: #7eb3f0;
            font-size: 14px;
            margin-bottom: 12px;
            padding-bottom: 6px;
            border-bottom: 1px solid #1e3a5f;
        }}

        /* ════════════════════════════════════
           상단 3패널: 수원시 분석 대시보드
           ════════════════════════════════════ */
        .top-dashboard {{
            display: grid;
            grid-template-columns: 220px 1fr 220px;
            gap: 0;
            background: #0f1c2e;
            border: 1px solid #1e3a5f;
            border-radius: 10px;
            margin-bottom: 16px;
            overflow: hidden;
        }}

        /* 패널 공통 */
        .panel {{
            padding: 12px;
            border-right: 1px solid #1e3a5f;
        }}
        .panel:last-child {{ border-right: none; }}
        .panel-title {{
            color: #7eb3f0;
            font-size: 13px;
            font-weight: bold;
            margin-bottom: 10px;
            padding-bottom: 6px;
            border-bottom: 1px solid #1e3a5f;
        }}

        /* 왼쪽 패널: 수원시 구별 */
        .panel-left {{ background: #0f1c2e; }}

        .gu-item {{ margin-bottom: 6px; }}
        .gu-header {{
            display: flex;
            align-items: center;
            gap: 6px;
            cursor: pointer;
            padding: 4px 6px;
            border-radius: 4px;
            transition: background 0.2s;
        }}
        .gu-header:hover {{ background: #1a2d47; }}
        .gu-arrow {{
            color: #ffa040;
            font-size: 10px;
            transition: transform 0.2s;
            display: inline-block;
        }}
        .gu-arrow.open {{ transform: rotate(90deg); }}
        .gu-name {{ color: #e2e8f0; font-size: 12px; font-weight: bold; }}
        .gu-detail {{
            list-style: none;
            padding: 4px 6px 4px 18px;
            font-size: 10px;
            color: #8a9bb0;
            line-height: 1.8;
        }}
        .gu-detail li::before {{ content: "· "; color: #4a7fb5; }}

        /* 중간 패널: 키워드 차트 */
        .panel-center {{ background: #0f1c2e; }}
        .chart-legend {{
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-bottom: 8px;
            font-size: 10px;
        }}
        .legend-dot {{
            width: 8px; height: 8px;
            border-radius: 50%;
            display: inline-block;
            margin-right: 3px;
            vertical-align: middle;
        }}

        /* 오른쪽 패널: 전문가 채널 브리핑 */
        .panel-right {{ background: #0f1c2e; }}

        .briefing-section-label {{
            font-size: 10px;
            color: #5a7a9a;
            margin: 8px 0 4px 0;
            font-weight: bold;
        }}
        .briefing-section-label:first-child {{ margin-top: 0; }}

        .briefing-card {{
            display: flex;
            align-items: center;
            gap: 7px;
            background: #13263d;
            border-radius: 5px;
            padding: 6px 7px;
            margin-bottom: 5px;
            border: 1px solid #1e3a5f;
        }}
        .briefing-yt-icon {{
            width: 28px; height: 28px;
            background: #cc0000;
            border-radius: 5px;
            display: flex;
            align-items: center;
            justify-content: center;
            color: #fff;
            font-size: 11px;
            flex-shrink: 0;
        }}
        .briefing-info {{ flex: 1; min-width: 0; }}
        .briefing-channels {{
            color: #e2e8f0;
            font-size: 10px;
            font-weight: bold;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        .briefing-desc {{
            color: #7a8fa8;
            font-size: 9px;
            margin-top: 1px;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        .briefing-duration {{
            color: #4a7fb5;
            font-size: 9px;
            flex-shrink: 0;
        }}

        /* ══ 유튜브 탭 ══ */
        .tab {{
            overflow: hidden;
            border-bottom: 2px solid #1e3a5f;
            margin-bottom: 12px;
        }}
        .tab button {{
            background: none;
            float: left;
            border: none;
            outline: none;
            cursor: pointer;
            padding: 8px 14px;
            font-size: 12px;
            color: #7a8fa8;
            font-weight: bold;
            transition: color 0.2s;
        }}
        .tab button:hover {{ color: #7eb3f0; }}
        .tab button.active {{
            color: #7eb3f0;
            border-bottom: 2px solid #4a9eff;
        }}
        .video-grid {{
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 8px;
        }}
        .video-item {{
            background: #0f1c2e;
            padding: 6px;
            border-radius: 6px;
            border: 1px solid #1e3a5f;
            font-size: 11px;
            color: #8a9bb0;
        }}

        /* ══ 뉴스 테이블 ══ */
        table {{ width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 12px; }}
        th {{
            background: #0f1c2e;
            color: #7eb3f0;
            padding: 10px 8px;
            text-align: left;
            border-bottom: 1px solid #1e3a5f;
        }}
        td {{
            padding: 8px;
            border-bottom: 1px solid #1a2d47;
            color: #b0c4d8;
        }}
        tr:hover td {{ background: #1a2d47; }}
        td a {{ color: #7eb3f0; text-decoration: none; }}
        td a:hover {{ text-decoration: underline; }}

        .badge {{
            padding: 2px 7px;
            border-radius: 10px;
            font-size: 10px;
            font-weight: bold;
        }}
        .국내 {{ background: #1e4a8a; color: #7eb3f0; }}
        .국외 {{ background: #5a1a1a; color: #f08080; }}

        /* ══ 상단 바 ══ */
        .topbar {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 16px;
            padding-bottom: 12px;
            border-bottom: 1px solid #1e3a5f;
        }}
        .dl-btn {{
            background: #1a5a3a;
            color: #5adf9a;
            padding: 8px 18px;
            text-decoration: none;
            border-radius: 6px;
            font-size: 13px;
            font-weight: bold;
            border: 1px solid #2a7a5a;
        }}
    </style>
</head>
<body>
<div class="container">

    <!-- 상단 바 -->
    <div class="topbar">
        <h1>📈 통합 마켓 인사이트 대시보드</h1>
        <a href="{OUTPUT_XLSX}" class="dl-btn">⬇ 엑셀 저장</a>
    </div>

    <!-- ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
         상단 3패널: 수원시 구별 분석 대시보드
         (스크린샷과 동일한 레이아웃)
         ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ -->
    <div class="top-dashboard">

        <!-- ① 왼쪽: 수원시 구별 상황 유형 -->
        <div class="panel panel-left">
            <div class="panel-title">수원시 구별 중동 상황 및 분석</div>
            <div style="font-size:10px; color:#5a7a9a; margin-bottom:8px;">수원시 구별 상황 유형</div>
            {suwon_panel_html}
        </div>

        <!-- ② 중간: 주간 유형별 키워드 검색량 차트 -->
        <div class="panel panel-center">
            <div class="panel-title">주간 유형별 키워드 검색량</div>
            <div class="chart-legend">
                <span><span class="legend-dot" style="background:#4a9eff;"></span>에너지/금융(가스,LPG 등)</span>
                <span><span class="legend-dot" style="background:#ff6b35;"></span>시장/공급(금리, 환율)</span>
                <span><span class="legend-dot" style="background:#00cc66;"></span>생활/인프라(물가, 물류)</span>
                <span><span class="legend-dot" style="background:#ffcc00;"></span>정책대응(지원금, 보조금)</span>
            </div>
            <div style="position:relative; height:200px;">
                <canvas id="keywordChart"></canvas>
            </div>
        </div>

        <!-- ③ 오른쪽: 전문가 채널 브리핑 -->
        <div class="panel panel-right">
            <div class="panel-title">전문가 채널 브리핑</div>
            {briefing_panel_html}
        </div>

    </div><!-- /.top-dashboard -->

    <!-- ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
         2열 그리드: 유튜브 분석
         ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ -->
    <div class="grid-2col">
        <div class="card" style="grid-column: 1 / -1;">
            <h3>🎥 전문가 분석 (채널 분류)</h3>
            <div class="tab">{tab_buttons}</div>
            {tab_contents}
        </div>
    </div>

    <!-- ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
         뉴스 섹션
         ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━ -->
    <div class="card" style="margin-bottom:16px;">
        <h3>🔥 실시간 뉴스 Top 5</h3>
        <ul style="padding-left:18px; line-height:2; font-size:13px;">
            {top5_html}
        </ul>
    </div>

    <div class="card">
        <h3>📰 뉴스 타임라인</h3>
        <table>
            <thead>
                <tr>
                    <th>시간(KST)</th><th>지역</th><th>분류</th><th>언론사</th><th>제목</th>
                </tr>
            </thead>
            <tbody>{news_rows}</tbody>
        </table>
    </div>

</div><!-- /.container -->

<script>
/* ── 유튜브 탭 전환 ── */
function openYoutubeTab(evt, tabName) {{
    document.querySelectorAll(".tabcontent").forEach(el => el.style.display = "none");
    document.querySelectorAll(".tablinks").forEach(el => el.classList.remove("active"));
    document.getElementById(tabName).style.display = "block";
    evt.currentTarget.classList.add("active");
}}

/* ── 구별 아코디언 ── */
function toggleGu(headerEl) {{
    const detail = headerEl.nextElementSibling;
    const arrow  = headerEl.querySelector(".gu-arrow");
    const isOpen = detail.style.display !== "none";
    detail.style.display = isOpen ? "none" : "block";
    arrow.classList.toggle("open", !isOpen);
}}

/* ── 키워드 검색량 꺾은선 차트 ── */
const kwCtx = document.getElementById('keywordChart').getContext('2d');
new Chart(kwCtx, {{
    type: 'line',
    data: {{
        labels: {json.dumps(chart_labels)},
        datasets: [
            {{
                label: '에너지/금융(가스,LPG 등)',
                data: [520, 480, 510, 490, 530, 500, 510],
                borderColor: '#4a9eff',
                backgroundColor: 'rgba(74,158,255,0.08)',
                tension: 0.4, fill: true, pointRadius: 3
            }},
            {{
                label: '시장/공급(금리, 환율)',
                data: [300, 320, 310, 290, 310, 300, 290],
                borderColor: '#ff6b35',
                backgroundColor: 'rgba(255,107,53,0.08)',
                tension: 0.4, fill: true, pointRadius: 3
            }},
            {{
                label: '생활/인프라(물가, 물류)',
                data: [200, 190, 210, 200, 195, 205, 200],
                borderColor: '#00cc66',
                backgroundColor: 'rgba(0,204,102,0.08)',
                tension: 0.4, fill: true, pointRadius: 3
            }},
            {{
                label: '정책대응(지원금, 보조금)',
                data: [100, 120, 110, 130, 115, 125, 120],
                borderColor: '#ffcc00',
                backgroundColor: 'rgba(255,204,0,0.08)',
                tension: 0.4, fill: true, pointRadius: 3
            }}
        ]
    }},
    options: {{
        responsive: true,
        maintainAspectRatio: false,
        plugins: {{
            legend: {{ display: false }}
        }},
        scales: {{
            x: {{
                ticks: {{ color: '#7a8fa8', font: {{ size: 10 }} }},
                grid:  {{ color: '#1a2d47' }}
            }},
            y: {{
                ticks: {{ color: '#7a8fa8', font: {{ size: 10 }} }},
                grid:  {{ color: '#1a2d47' }}
            }}
        }}
    }}
}});
</script>
</body>
</html>"""

    os.makedirs(os.path.dirname(OUTPUT_HTML), exist_ok=True)
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"  ✅ HTML 저장 완료: {OUTPUT_HTML}")

# ─────────────────────────────────────
# [4] 실행부
# ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 50)
    print("통합 파이프라인 가동...")
    econ_data, econ_dates = get_econ_indicators()
    all_news = collect_news()
    save_all(all_news, econ_data, econ_dates)
    print("✅ 완료! market.html을 확인하세요.")
    print("=" * 50)
