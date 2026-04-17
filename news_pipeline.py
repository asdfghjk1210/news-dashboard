# news_pipeline.py
import sys
import os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if os.environ.get("ANTHROPIC_API_KEY", ""):
    import anthropic
else:
    anthropic = None
import feedparser
import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import urllib.parse
import time
import re

# ─────────────────────────────────────
# ★ 설정
# ─────────────────────────────────────
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
OUTPUT_XLSX = "news_result.xlsx"
OUTPUT_HTML = "docs/index.html"
DAYS_RANGE  = 30

KEYWORDS = {
    "관광 및 K-컬처": ["K-컬처", "K-culture", "K-푸드", "K-food", "K-POP", "한류"],
    "혁신산업":       ["바이오", "반도체", "AI", "인공지능"],
    "경제자유구역":   ["경제자유구역", "경자특구", "경자구역"],
    "수원시 관련":    ["수원시", "수원특례시"],
}

CATEGORY_COLORS = {
    "관광 및 K-컬처": {"bg": "#D6E4F0", "text": "#0C447C", "header": "#2E75B6"},
    "혁신산업":       {"bg": "#D5F5E3", "text": "#27500A", "header": "#2E8B57"},
    "경제자유구역":   {"bg": "#FEF9E7", "text": "#633806", "header": "#D4A017"},
    "수원시 관련":    {"bg": "#FDEDEC", "text": "#A32D2D", "header": "#C0392B"},
}

# ─────────────────────────────────────
# 1. Google News RSS 크롤링
# ─────────────────────────────────────
def get_google_news(keyword: str, category: str) -> list[dict]:
    encoded = urllib.parse.quote(keyword)
    url = f"https://news.google.com/rss/search?q={encoded}&hl=ko&gl=KR&ceid=KR:ko"
    feed = feedparser.parse(url)
    results = []
    cutoff = datetime.now(timezone.utc) - timedelta(days=DAYS_RANGE)
    for entry in feed.entries:
        try:
            pub = datetime(*entry.published_parsed[:6], tzinfo=timezone.utc)
        except Exception:
            continue
        if pub < cutoff:
            continue
        source = entry.get("source", {})
        results.append({
            "유형":     category,
            "키워드":   keyword,
            "기사제목": entry.get("title", ""),
            "기사링크": entry.get("link", ""),
            "보도일":   pub,
            "언론사":   source.get("title", "") if isinstance(source, dict) else "",
        })
    return results


def crawl_all() -> list[dict]:
    all_articles = []
    for category, kw_list in KEYWORDS.items():
        for kw in kw_list:
            print(f"  크롤링 중: [{category}] {kw}")
            try:
                articles = get_google_news(kw, category)
                all_articles.extend(articles)
                time.sleep(0.5)
            except Exception as e:
                print(f"    오류: {e}")
    seen = set()
    unique = []
    for a in all_articles:
        if a["기사링크"] not in seen:
            seen.add(a["기사링크"])
            unique.append(a)
    print(f"\n수집 완료: {len(unique)}건 (중복 제거 후)")
    return unique


# ─────────────────────────────────────
# 2. 본문 수집
# ─────────────────────────────────────
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ko-KR,ko;q=0.9",
}

SELECTORS = [
    "article p", "#articleBody p", "#articeBody p",
    ".article_body p", ".article-body p", ".article_txt p",
    ".news_text p", ".news-content p", ".content p",
    ".story-news p", ".main_txt p", ".text p",
    "#newsct_article p", ".go_trans p", "p",
]

def resolve_google_news_url(url: str) -> str:
    try:
        clean = re.sub(r"\?.*$", "", url)
        clean = clean.replace("/rss/articles/", "/articles/")
        session = requests.Session()
        session.headers.update(HEADERS)
        resp = session.get(clean, timeout=15, allow_redirects=True)
        final_url = resp.url
        if final_url and "google.com" not in final_url:
            return final_url
        html = resp.text
        for pattern in [
            r'data-n-au="([^"]+)"',
            r'window\.location\.href\s*=\s*["\']([^"\']+)["\']',
            r'window\.location\s*=\s*["\']([^"\']+)["\']',
        ]:
            m = re.search(pattern, html)
            if m and "google.com" not in m.group(1):
                return m.group(1)
    except Exception:
        pass
    try:
        from newspaper import Article
        article = Article(url, language="ko")
        article.download()
        article.parse()
        if article.canonical_link and "google.com" not in article.canonical_link:
            return article.canonical_link
    except Exception:
        pass
    return url


def get_article_content(url: str) -> str:
    try:
        real_url = resolve_google_news_url(url)
        r = requests.get(real_url, headers=HEADERS, timeout=10)
        r.encoding = r.apparent_encoding
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "header", "footer", "aside"]):
            tag.decompose()
        for selector in SELECTORS:
            nodes = soup.select(selector)
            text = " ".join(n.get_text(strip=True) for n in nodes)
            text = re.sub(r"\s+", " ", text).strip()
            if len(text) >= 100:
                return text
        try:
            from newspaper import Article
            article = Article(real_url, language="ko")
            article.download()
            article.parse()
            if article.text and len(article.text) >= 100:
                return article.text
        except Exception:
            pass
        return "본문 수집 실패"
    except Exception:
        return "본문 수집 실패"


def _fallback_summary(text: str, max_chars: int = 200) -> str:
    sentences = re.split(r"(?<=[.!?。])\s+", text)
    summary = ""
    for s in sentences:
        if len(summary) + len(s) <= max_chars:
            summary += s + " "
        else:
            break
    return summary.strip() if summary else text[:max_chars] + "..."


def make_summary(text: str, title: str = "", max_chars: int = 200) -> str:
    source = text if (text and text != "본문 수집 실패") else title
    if not source:
        return ""
    if not ANTHROPIC_API_KEY:
        return _fallback_summary(source, max_chars)
    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        prompt = f"다음 내용을 2~3문장으로 한국어 요약해줘. 요약문만 출력해.\n\n{source[:3000]}"
        msg = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        return msg.content[0].text.strip()
    except Exception:
        return _fallback_summary(source, max_chars)


# ─────────────────────────────────────
# 3. 엑셀 저장
# ─────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CATEGORY_FILLS = {
    "관광 및 K-컬처": "D6E4F0",
    "혁신산업":       "D5F5E3",
    "경제자유구역":   "FEF9E7",
    "수원시 관련":    "FDEDEC",
}

def save_excel(articles: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "뉴스모니터링"
    cols = ["일자", "일시", "유형", "언론사", "기사제목", "요약", "기사본문", "URL"]
    ws.append(cols)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for a in articles:
        pub: datetime = a["보도일"]
        row = [
            pub.strftime("%Y-%m-%d"),
            pub.strftime("%H:%M"),
            a.get("유형", ""),
            a.get("언론사", ""),
            a.get("기사제목", ""),
            a.get("요약", ""),
            a.get("본문", ""),
            a.get("기사링크", ""),
        ]
        ws.append(row)
        fill_color = CATEGORY_FILLS.get(a.get("유형", ""), "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    col_widths = [12, 8, 14, 16, 50, 60, 80, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    wb.save(path)
    print(f"엑셀 저장 완료: {path}")


# ─────────────────────────────────────
# 4. HTML 대시보드 저장
# ─────────────────────────────────────
def save_html(articles: list[dict], path: str):
    import json
    from collections import Counter

    now_str = datetime.now().strftime("%Y년 %m월 %d일")
    total = len(articles)

    # ── 통계 카드 ──
    counts = {cat: sum(1 for a in articles if a.get("유형") == cat) for cat in KEYWORDS}

    stat_cards = ""
    for cat, cnt in counts.items():
        c = CATEGORY_COLORS.get(cat, {"bg": "#f0f0f0", "text": "#333", "header": "#666"})
        stat_cards += f"""
        <div style="background:{c['bg']};border-radius:10px;padding:16px 20px;min-width:140px;flex:1;">
          <div style="font-size:12px;color:{c['text']};margin-bottom:6px;font-weight:500;">{cat}</div>
          <div style="font-size:28px;font-weight:700;color:{c['header']};">{cnt}</div>
          <div style="font-size:11px;color:{c['text']};margin-top:4px;">건</div>
        </div>"""

    # ── 차트 데이터 계산 ──
    cat_header_colors = {cat: CATEGORY_COLORS[cat]["header"] for cat in KEYWORDS}

    # 1. 날짜별 유형별 추이
    date_cat: dict = {}
    for a in articles:
        d = a["보도일"].strftime("%Y-%m-%d")
        cat = a.get("유형", "")
        date_cat.setdefault(d, {})
        date_cat[d][cat] = date_cat[d].get(cat, 0) + 1
    sorted_dates = sorted(date_cat.keys())
    line_datasets = [
        {
            "label": cat,
            "data": [date_cat.get(d, {}).get(cat, 0) for d in sorted_dates],
            "borderColor": cat_header_colors[cat],
            "backgroundColor": cat_header_colors[cat],
            "tension": 0.3,
            "fill": False,
            "pointRadius": 3,
        }
        for cat in KEYWORDS
    ]

    # 2. 유형별 도넛
    donut_labels = list(counts.keys())
    donut_data   = list(counts.values())
    donut_colors = [cat_header_colors[c] for c in donut_labels]

    # 3. 언론사 TOP 10
    press_counter = Counter(a.get("언론사", "") for a in articles if a.get("언론사", "").strip())
    top_press     = press_counter.most_common(10)
    press_labels  = [p[0] for p in top_press]
    press_data    = [p[1] for p in top_press]

    # 4. 제목 키워드 빈도 TOP 10
    stop_words = {
        "및", "등", "위해", "대한", "통해", "관련", "이후", "이번", "지난", "오는",
        "대해", "있는", "하는", "으로", "에서", "로서", "부터", "까지", "하여",
        "이를", "또한", "이에", "따라", "위한", "한국", "위해", "있어", "통한",
        "대한민국", "있다", "않는", "없는", "되는", "하고", "하며",
    }
    word_counter = Counter()
    for a in articles:
        for w in re.findall(r"[가-힣]{2,}", a.get("기사제목", "")):
            if w not in stop_words:
                word_counter[w] += 1
    top_words  = word_counter.most_common(10)
    kw_labels  = [w[0] for w in top_words]
    kw_data    = [w[1] for w in top_words]

    # JSON 직렬화
    j_dates          = json.dumps(sorted_dates,  ensure_ascii=False)
    j_line_datasets  = json.dumps(line_datasets, ensure_ascii=False)
    j_donut_labels   = json.dumps(donut_labels,  ensure_ascii=False)
    j_donut_data     = json.dumps(donut_data)
    j_donut_colors   = json.dumps(donut_colors)
    j_press_labels   = json.dumps(press_labels,  ensure_ascii=False)
    j_press_data     = json.dumps(press_data)
    j_kw_labels      = json.dumps(kw_labels,     ensure_ascii=False)
    j_kw_data        = json.dumps(kw_data)

    # ── 기사 행 ──
    rows = ""
    for a in articles:
        pub: datetime = a["보도일"]
        cat = a.get("유형", "")
        c = CATEGORY_COLORS.get(cat, {"bg": "#ffffff", "text": "#333", "header": "#333"})
        title   = a.get("기사제목", "").replace("<", "&lt;").replace(">", "&gt;")
        summary = a.get("요약",     "").replace("<", "&lt;").replace(">", "&gt;")
        press   = a.get("언론사",   "").replace("<", "&lt;").replace(">", "&gt;")
        raw_url = a.get("기사링크", "")
        m = re.search(r"/articles/([^?&]+)", raw_url)
        url = f"https://news.google.com/articles/{m.group(1)}" if m else raw_url
        rows += f"""
        <tr style="background:{c['bg']};">
          <td style="white-space:nowrap;color:#555;font-size:12px;">{pub.strftime("%Y-%m-%d")}</td>
          <td style="white-space:nowrap;color:#555;font-size:12px;">{pub.strftime("%H:%M")}</td>
          <td><span style="background:{c['header']};color:#fff;font-size:11px;padding:3px 8px;border-radius:12px;white-space:nowrap;">{cat}</span></td>
          <td style="font-size:12px;color:#555;white-space:nowrap;">{press}</td>
          <td style="font-weight:500;"><a href="{url}" target="_blank" style="color:{c['header']};text-decoration:none;">{title}</a></td>
          <td style="font-size:12px;color:#666;line-height:1.5;">{summary}</td>
          <td style="text-align:center;"><a href="{url}" target="_blank" style="background:{c['header']};color:#fff;font-size:11px;padding:4px 10px;border-radius:6px;text-decoration:none;white-space:nowrap;">원문 보기</a></td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>뉴스 모니터링 리포트</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Malgun Gothic', Arial, sans-serif; background: #f5f6fa; color: #333; }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 24px; }}
  .header {{ background: #1F4E79; color: #fff; border-radius: 12px; padding: 24px 28px; margin-bottom: 20px; }}
  .header h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 6px; }}
  .header p {{ font-size: 13px; opacity: 0.8; }}
  .stats {{ display: flex; gap: 14px; margin-bottom: 20px; flex-wrap: wrap; }}
  .total-card {{ background: #fff; border-radius: 10px; padding: 16px 20px; min-width: 140px; flex: 1; border: 1px solid #e0e0e0; }}
  .total-card .label {{ font-size: 12px; color: #888; margin-bottom: 6px; }}
  .total-card .value {{ font-size: 28px; font-weight: 700; color: #1F4E79; }}
  .charts-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px; }}
  .chart-card {{ background: #fff; border-radius: 12px; border: 1px solid #e0e0e0; padding: 20px; }}
  .chart-card h3 {{ font-size: 13px; font-weight: 600; color: #555; margin-bottom: 14px; }}
  .chart-card canvas {{ max-height: 260px; }}
  .table-wrap {{ background: #fff; border-radius: 12px; overflow: hidden; border: 1px solid #e0e0e0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  thead tr {{ background: #1F4E79; }}
  thead th {{ color: #fff; font-weight: 600; padding: 12px 14px; text-align: left; white-space: nowrap; }}
  tbody tr {{ border-bottom: 1px solid #eee; }}
  tbody tr:hover {{ filter: brightness(0.97); }}
  td {{ padding: 10px 14px; vertical-align: top; }}
  a:hover {{ text-decoration: underline !important; }}
</style>
</head>
<body>
<div class="container">
  <div class="header">
    <h1>뉴스 모니터링 리포트</h1>
    <p>{now_str} 기준 · 최근 {DAYS_RANGE}일 수집</p>
  </div>
  <div class="stats">
    <div class="total-card">
      <div class="label">전체 기사</div>
      <div class="value">{total}</div>
    </div>
    {stat_cards}
  </div>

  <div class="charts-grid">
    <div class="chart-card">
      <h3>날짜별 유형별 기사 추이</h3>
      <canvas id="chartLine"></canvas>
    </div>
    <div class="chart-card">
      <h3>유형별 비율</h3>
      <canvas id="chartDonut"></canvas>
    </div>
    <div class="chart-card">
      <h3>언론사별 기사 수 (TOP 10)</h3>
      <canvas id="chartPress"></canvas>
    </div>
    <div class="chart-card">
      <h3>키워드 빈도 (TOP 10)</h3>
      <canvas id="chartKeyword"></canvas>
    </div>
  </div>

  <div class="table-wrap">
    <table>
      <thead>
        <tr>
          <th>일자</th>
          <th>일시</th>
          <th>유형</th>
          <th>언론사</th>
          <th>기사제목</th>
          <th>요약</th>
          <th>원문</th>
        </tr>
      </thead>
      <tbody>
        {rows}
      </tbody>
    </table>
  </div>
</div>

<script>
(function() {{
  // 1. 날짜별 유형별 꺾은선 차트
  new Chart(document.getElementById('chartLine'), {{
    type: 'line',
    data: {{
      labels: {j_dates},
      datasets: {j_line_datasets}
    }},
    options: {{
      responsive: true,
      plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }} }} }} }},
      scales: {{
        x: {{ ticks: {{ font: {{ size: 10 }}, maxRotation: 45 }} }},
        y: {{ beginAtZero: true, ticks: {{ stepSize: 1, font: {{ size: 11 }} }} }}
      }}
    }}
  }});

  // 2. 유형별 도넛 차트
  new Chart(document.getElementById('chartDonut'), {{
    type: 'doughnut',
    data: {{
      labels: {j_donut_labels},
      datasets: [{{ data: {j_donut_data}, backgroundColor: {j_donut_colors}, borderWidth: 2 }}]
    }},
    options: {{
      responsive: true,
      plugins: {{
        legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }} }} }},
        tooltip: {{ callbacks: {{ label: function(ctx) {{ return ctx.label + ': ' + ctx.parsed + '건'; }} }} }}
      }}
    }}
  }});

  // 3. 언론사 가로 막대 차트
  new Chart(document.getElementById('chartPress'), {{
    type: 'bar',
    data: {{
      labels: {j_press_labels},
      datasets: [{{ label: '기사 수', data: {j_press_data}, backgroundColor: '#2E75B6', borderRadius: 4 }}]
    }},
    options: {{
      indexAxis: 'y',
      responsive: true,
      plugins: {{ legend: {{ display: false }} }},
      scales: {{
        x: {{ beginAtZero: true, ticks: {{ stepSize: 1, font: {{ size: 11 }} }} }},
        y: {{ ticks: {{ font: {{ size: 11 }} }} }}
      }}
    }}
  }});

  // 4. 키워드 빈도 막대 차트
  new Chart(document.getElementById('chartKeyword'), {{
    type: 'bar',
    data: {{
      labels: {j_kw_labels},
      datasets: [{{ label: '빈도', data: {j_kw_data}, backgroundColor: '#2E8B57', borderRadius: 4 }}]
    }},
    options: {{
      responsive: true,
      plugins: {{ legend: {{ display: false }} }},
      scales: {{
        x: {{ ticks: {{ font: {{ size: 11 }} }} }},
        y: {{ beginAtZero: true, ticks: {{ stepSize: 1, font: {{ size: 11 }} }} }}
      }}
    }}
  }});
}})();
</script>
</body>
</html>"""

    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML 저장 완료: {path}")


# ─────────────────────────────────────
# 5. 완료
# ─────────────────────────────────────


# ─────────────────────────────────────
# 메인 실행
# ─────────────────────────────────────
if __name__ == "__main__":
    print("=" * 50)
    print("뉴스 모니터링 파이프라인 시작")
    print("=" * 50)

    print("\n[1/5] Google News RSS 크롤링...")
    articles = crawl_all()

    print("\n[2/5] 기사 본문 + 요약 수집 중...")
    for i, a in enumerate(articles):
        print(f"  ({i+1}/{len(articles)}) {a['기사제목'][:30]}...", flush=True)
        content = get_article_content(a["기사링크"])
        a["본문"] = content
        a["요약"] = make_summary(content, title=a.get("기사제목", ""))
        time.sleep(0.3)

    success = sum(1 for a in articles if a["본문"] != "본문 수집 실패")
    print(f"  본문 수집 성공률: {success}/{len(articles)} ({round(success/len(articles)*100,1)}%)")

    print("\n[3/5] 엑셀 파일 생성 중...")
    save_excel(articles, OUTPUT_XLSX)

    print("\n[4/5] HTML 대시보드 생성 중...")
    save_html(articles, OUTPUT_HTML)

    print("\n" + "=" * 50)
    print("전체 파이프라인 완료!")
    print("=" * 50)